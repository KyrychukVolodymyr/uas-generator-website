from openpyxl import load_workbook

def _tt_num(value):
    try:
        if value in (None, ""):
            return 0
        return float(value)
    except Exception:
        return 0

def _tt_recalc_sheet(ws):
    """
    Keep TT formulas alive.

    The program still fills CSV-derived input cells in B:H and I:J.
    This function must NOT replace formula cells with hardcoded numbers.
    K12:K26 and K27 stay as Excel formulas so manual edits recalculate.
    """
    for row in range(12, 27):
        ws[f"K{row}"] = f"=SUM(B{row}:H{row})*I{row}*J{row}"

    ws["K27"] = "=SUM(K12:K26)/60"

    try:
        wb = ws.parent
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

def _rows_blob(rows):
    parts = []
    for row in rows:
        if isinstance(row, (list, tuple)):
            joined = " | ".join(str(c).strip() for c in row if str(c).strip())
        else:
            joined = str(row).strip()
        if joined:
            parts.append(joined.lower())
    return "\n".join(parts)

def _row_has_minutes(ws, row_num):
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        v = ws[f"{col}{row_num}"].value
        try:
            if v not in (None, "") and float(v) > 0:
                return True
        except Exception:
            pass
    return False

def apply_tt_schedule_override_v2(xlsx_path, days_per_week, hours_per_day, rows, service_flag=False):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    if hours_per_day is not None:
        ws["B32"] = hours_per_day

    if days_per_week is not None:
        ws["K32"] = days_per_week

    footer_days = ws["K32"].value
    footer_hours = ws["B32"].value

    if footer_days not in (None, "") and footer_hours not in (None, ""):
        footer = f"PCA {footer_days} D x {footer_hours} H"
    elif footer_days not in (None, ""):
        footer = f"PCA {footer_days} D"
    else:
        footer = "PCA"

    if service_flag:
        ws["A36"] = footer + "\nWould like to discuss service hrs with CM."
    else:
        ws["A36"] = footer

    blob = _rows_blob(rows)
    is_scic = "significant change in status reassessment" in blob

    if is_scic:
        scic_rows = {
            12: 7, 13: 7, 14: 3, 15: 7, 16: 2,
            17: 7, 18: 7, 19: 7, 20: 7, 21: 7,
            22: 7, 23: 7, 24: 7
        }
        for r, v in scic_rows.items():
            ws[f"J{r}"] = v
    else:
        if days_per_week is not None:
            for r in [12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24]:
                ws[f"J{r}"] = days_per_week
            ws["J14"] = 3
            ws["J16"] = 2

    laundry_continent_active = _row_has_minutes(ws, 25)
    laundry_incontinent_active = _row_has_minutes(ws, 26)

    if laundry_incontinent_active:
        ws["J25"] = None
        ws["J26"] = 3
    elif laundry_continent_active:
        ws["J25"] = 2
        ws["J26"] = None
    else:
        ws["J25"] = None
        ws["J26"] = None

    _tt_recalc_sheet(ws)
    wb.save(xlsx_path)
