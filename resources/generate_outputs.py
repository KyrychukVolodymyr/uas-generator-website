
def restore_excel_images_only(template_path, output_path):
    import zipfile
    import shutil
    import tempfile
    from pathlib import Path

    temp_dir = Path(tempfile.mkdtemp())

    with zipfile.ZipFile(template_path, 'r') as z:
        z.extractall(temp_dir / "template")

    with zipfile.ZipFile(output_path, 'r') as z:
        z.extractall(temp_dir / "output")

    src_media = temp_dir / "template/xl/media"
    dst_media = temp_dir / "output/xl/media"

    src_draw = temp_dir / "template/xl/drawings"
    dst_draw = temp_dir / "output/xl/drawings"

    if src_media.exists():
        shutil.rmtree(dst_media, ignore_errors=True)
        shutil.copytree(src_media, dst_media)

    if src_draw.exists():
        shutil.rmtree(dst_draw, ignore_errors=True)
        shutil.copytree(src_draw, dst_draw)

    new_file = output_path.with_suffix(".fixed.xlsx")

    with zipfile.ZipFile(new_file, 'w', zipfile.ZIP_DEFLATED) as z:
        for file in (temp_dir / "output").rglob("*"):
            z.write(file, file.relative_to(temp_dir / "output"))

    new_file.replace(output_path)




import argparse
from pathlib import Path


def _fra_parse_date_safe(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _fra_exact_age_from_patient(patient):
    header = patient.get("header", {})
    dob = _fra_parse_date_safe(header.get("dob"))
    assessment = _fra_parse_date_safe(header.get("assessment_date")) or _fra_parse_date_safe(header.get("date_of_assessment"))
    if not dob or not assessment:
        return None
    years = assessment.year - dob.year
    if (assessment.month, assessment.day) < (dob.month, dob.day):
        years -= 1
    return years


def _fra_has_cognitive_keyword(patient):
    chunks = []

    header = patient.get("header", {})
    for v in header.values():
        if v is not None:
            chunks.append(str(v))

    diagnoses = patient.get("diagnoses", [])
    if isinstance(diagnoses, list):
        chunks.extend(str(x) for x in diagnoses if x is not None)

    fra_reasons = patient.get("fra_reasons", [])
    if isinstance(fra_reasons, list):
        chunks.extend(str(x) for x in fra_reasons if x is not None)

    raw_text = patient.get("raw_text")
    if raw_text:
        chunks.append(str(raw_text))

    blob = " | ".join(chunks).lower()
    return "cognitive" in blob



def _fra_parse_date(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _fra_extract_first_date_matching(rows, labels):
    labels = [x.lower() for x in labels]
    for row in rows:
        vals = ["" if v is None else str(v).strip() for v in row]
        for i, cell in enumerate(vals):
            low = cell.lower()
            if any(lbl in low for lbl in labels):
                d = _fra_parse_date(cell)
                if d:
                    return d
                for j in range(i + 1, min(i + 6, len(vals))):
                    d = _fra_parse_date(vals[j])
                    if d:
                        return d
    return None

def _fra_exact_age_years_from_rows(rows):
    dob = _fra_extract_first_date_matching(rows, ["date of birth", "dob", "birth date"])
    assessment = _fra_extract_first_date_matching(rows, ["assessment date", "date of assessment", "reference date"])
    if not dob or not assessment:
        return None
    years = assessment.year - dob.year
    if (assessment.month, assessment.day) < (dob.month, dob.day):
        years -= 1
    return years

def _fra_rows_have_cognitive(rows):
    for row in rows:
        blob = " | ".join("" if v is None else str(v) for v in row).lower()
        if "cognitive" in blob:
            return True
    return False


import openpyxl
import pandas as pd
from pypdf import PdfReader, PdfWriter

from uas_automation import (
    parse_header,
    parse_tasks,
    parse_incontinence,
    collect_positive_diagnoses,
    collect_med_rows,
    build_cdpas,
    build_fra,
)
from commercial_profile import get_assessor_display_name, get_assessor_plain_name
from io import BytesIO
from reportlab.pdfgen import canvas

UMBRELLA_DIR = Path.home() / "Documents" / "Anthem Blue Cross Blue Shield Assessment Documentation"

TT_ROWS = {
    "cognition": 12,
    "meal preparation": 13,
    "ordinary housework": 14,
    "managing medications": 15,
    "shopping": 16,
    "bathing": 17,
    "personal hygiene": 18,
    "dressing upper body": 19,
    "dressing lower body": 20,
    "locomotion": 21,
    "toilet use": 22,
    "transfer toilet": 23,
    "eating": 24,
}

TT_SCORE_TO_COL = {
    3: "E",
    4: "F",
    5: "G",
    6: "H",
}

TT_MINUTE_GRID = {
    "cognition": {3: 0, 4: 5, 5: 10, 6: 20},
    "meal preparation": {3: 15, 4: 20, 5: 20, 6: 25},
    "ordinary housework": {3: 20, 4: 30, 5: 45, 6: 60},
    "managing medications": {3: 5, 4: 5, 5: 5, 6: 10},
    "shopping": {3: 60, 4: 60, 5: 60, 6: 60},
    "bathing": {3: 10, 4: 20, 5: 30, 6: 30},
    "personal hygiene": {3: 5, 4: 5, 5: 10, 6: 10},
    "dressing upper body": {3: 5, 4: 5, 5: 10, 6: 15},
    "dressing lower body": {3: 5, 4: 5, 5: 10, 6: 15},
    "locomotion": {3: 5, 4: 5, 5: 5, 6: 5},
    "toilet use": {3: 10, 4: 10, 5: 15, 6: 15},
    "transfer toilet": {3: 5, 4: 5, 5: 10, 6: 20},
    "eating": {3: 15, 4: 15, 5: 20, 6: 30},
}

TT_DAILY_FREQUENCY = {
    "cognition": 2,
    "meal preparation": 3,
    "ordinary housework": 1,
    "managing medications": 3,
    "shopping": 1,
    "bathing": 1,
    "personal hygiene": 2,
    "dressing upper body": 2,
    "dressing lower body": 2,
    "locomotion": 4,
    "toilet use": 5,
    "transfer toilet": 5,
    "eating": 3,
    "laundry_continent": 1,
    "laundry_incontinent": 1,
}

TT_DAYS_PER_WEEK = {
    "cognition": 7,
    "meal preparation": 7,
    "ordinary housework": 3,
    "managing medications": 7,
    "shopping": 2,
    "bathing": 7,
    "personal hygiene": 7,
    "dressing upper body": 7,
    "dressing lower body": 7,
    "locomotion": 7,
    "toilet use": 7,
    "transfer toilet": 7,
    "eating": 7,
    "laundry_continent": 2,
    "laundry_incontinent": 3,
}



def _tt_num(value):
    try:
        if value is None or value == "":
            return 0
        return float(value)
    except Exception:
        return 0

def _tt_recalc_sheet(ws):
    """
    Keep TT formulas alive.

    The program still fills CSV-derived input cells in B:H and I:J.
    This function must NOT replace formula cells with hardcoded numbers.
    K12:K26 and K27 stay as Excel formulas so manual edits recalculate.
    """
    for row in range(12, 27):
        ws[f"K{row}"] = f"=SUM(B{row}:H{row})*I{row}*J{row}"

    ws["K27"] = "=SUM(K12:K26)/60"

    try:
        wb = ws.parent
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass



def _tt_cognition_is_modified_independent(patient):
    df = patient.get("df")
    if df is None:
        return False
    try:
        for _, row in df.iterrows():
            vals = ["" if x is None else str(x).strip().lower() for x in row.tolist()]
            blob = " | ".join(vals)
            if "making decisions regarding tasks of daily life" in blob and "modified independence" in blob:
                return True
    except Exception:
        return False
    return False


def load_patient(csv_path: str):
    df = pd.read_csv(csv_path, header=None)

    header = parse_header(df)
    tasks = parse_tasks(df)
    incontinent, bladder, bowel = parse_incontinence(df)
    positive_diagnoses = collect_positive_diagnoses(df)
    med_rows = collect_med_rows(df)
    cdpas = build_cdpas(df, med_rows)
    fra_score, fra_reasons, med_count, diag_count = build_fra(
        header, incontinent, positive_diagnoses, med_rows, df
    )

    return {
        "df": df,
        "header": header,
        "tasks": tasks,
        "incontinent": incontinent,
        "bladder": bladder,
        "bowel": bowel,
        "positive_diagnoses": positive_diagnoses,
        "med_rows": med_rows,
        "cdpas": cdpas,
        "fra_score": fra_score,
        "fra_reasons": fra_reasons,
        "med_count": med_count,
        "diag_count": diag_count,
    }


def build_base_name(patient, case_id: str) -> str:
    cin = patient["header"]["cin"]
    name_file = patient["header"]["patient_name_file"]
    return f"ANTHEM_MLTC-{cin}-{name_file}-{case_id}"


def build_patient_folder(patient) -> Path:
    name_file = patient["header"]["patient_name_file"]
    assessment_date = patient["header"]["assessment_date"] or "NO-DATE"
    safe_date = assessment_date.replace("/", "-")
    return UMBRELLA_DIR / f"{name_file} - {safe_date}"


def tt_display_name(patient) -> str:
    return patient["header"]["patient_name_file"]


def fill_tt(patient, template_path: str, output_path: Path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Header
    ws["A10"] = f"Member:  {tt_display_name(patient)}"
    ws["B10"] = f"Member DOB:  {patient['header']['dob']}"
    ws["A44"] = get_assessor_display_name()
    # keep assessment date from CSV, not today's date, to match patient folder / source assessment
    ws["B44"] = f"Date: {patient['header']['assessment_date']}"
# Task rows
    for task, row in TT_ROWS.items():
        # clear score columns only
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"] = None

        # always set frequency/days
        ws[f"I{row}"] = TT_DAILY_FREQUENCY[task]
        ws[f"J{row}"] = TT_DAYS_PER_WEEK[task]

        csv_score = patient["tasks"].get(task)
        if csv_score is None:
            continue

        if task == "cognition" and csv_score == 3:
            ws["E12"] = "0"
            ws["E12"].number_format = "@"
            continue

        # cognition already mapped by parser:
        # modified independent -> 3
        # minimally impaired -> 4
        # moderately impaired -> 5
        # severely impaired / no discernible consciousness -> 6
        if csv_score not in TT_SCORE_TO_COL:
            continue

        minute_value = TT_MINUTE_GRID[task].get(csv_score)
        if minute_value is None:
            continue

        col_letter = TT_SCORE_TO_COL[csv_score]
        ws[f"{col_letter}{row}"] = minute_value

    # Laundry rows
    for row in [25, 26]:
        for col in ["B", "C", "D", "E", "F", "G", "H"]:
            ws[f"{col}{row}"] = None

    ws["I25"] = None
    ws["J25"] = None
    ws["I26"] = None
    ws["J26"] = None

    if patient["incontinent"]:
        ws["H26"] = 60
        ws["I26"] = TT_DAILY_FREQUENCY["laundry_incontinent"]
        ws["J26"] = TT_DAYS_PER_WEEK["laundry_incontinent"]
    else:
        ws["H25"] = 60
        ws["I25"] = TT_DAILY_FREQUENCY["laundry_continent"]
        ws["J25"] = TT_DAYS_PER_WEEK["laundry_continent"]
    _tt_recalc_sheet(ws)

    wb.save(output_path)
    restore_excel_images_only(template_path, output_path)



def fill_cdpas(patient, template_path: str, output_path: Path):
    print('V020Q_GENERATE_OUTPUTS_FILL_CDPAS_ENTERED', flush=True)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    ws["B1"] = patient["header"]["patient_name_file"]
    ws["C1"] = f"DOB:  {patient['header']['dob']}"
    ws["E1"] = f"CIN: {patient['header']['cin']}"
    ws["E2"] = f"Date of Plan: {patient['header']['today_date']}"
    ws["E4"] = f"RN Signature: {get_assessor_display_name()}"

    for cell in ["G11", "G30", "G35"]:
        ws[cell] = None

    if "Insulin" in patient["cdpas"]:
        ws["G11"] = "X"

    if "Oxygen" in patient["cdpas"]:
        ws["G30"] = "X"

    if "Wound Care" in patient["cdpas"]:
        ws["G35"] = "X"

    wb.save(output_path)
    restore_excel_images_only(template_path, output_path)


def fill_fra(patient, template_path: str, output_path: Path, case_id: str):
    reader = PdfReader(template_path)
    fra_exact_age = _fra_exact_age_from_patient(patient)
    fra_has_cognitive = _fra_has_cognitive_keyword(patient)
    writer = PdfWriter()
    writer.clone_document_from_reader(reader)

    fra_reasons = patient.get("fra_reasons", [])
    reasons_text = " | ".join(str(x) for x in fra_reasons if x is not None)
    reasons_lower = reasons_text.lower()

    age_65 = 1 if (fra_exact_age is not None and fra_exact_age >= 65) else 0
    diagnosis_3_plus = 1 if "3+ diagnoses" in reasons_lower else 0
    falls = 1 if "falls" in reasons_lower else 0
    incontinence = 1 if "incontinence" in reasons_lower else 0
    visual_impairment = 1 if ("visual" in reasons_lower or "vision" in reasons_lower or "blind" in reasons_lower) else 0
    impaired_functional_mobility = 1 if ("functional mobility" in reasons_lower or "mobility" in reasons_lower or "gait" in reasons_lower or "transfer" in reasons_lower) else 0
    environmental_hazards = 1 if ("environment" in reasons_lower or "hazard" in reasons_lower or "clutter" in reasons_lower) else 0
    polypharmacy = 1 if "polypharmacy" in reasons_lower else 0
    pain = 1 if "pain" in reasons_lower else 0
    cognitive_impairment = 1 if (("cognitive impairment diagnosis" in reasons_lower) or fra_has_cognitive or ("cognitive" in reasons_lower) or ("dementia" in reasons_lower) or ("alzheimer" in reasons_lower)) else 0

    fra_total = (
        age_65
        + diagnosis_3_plus
        + falls
        + incontinence
        + visual_impairment
        + impaired_functional_mobility
        + environmental_hazards
        + polypharmacy
        + pain
        + cognitive_impairment
    )

    assessor_plain = get_assessor_plain_name()

    fields = {
        "Member Name": patient["header"]["patient_name_file"],
        "Member ID": str(case_id),
        "Date of Assessment": patient["header"]["assessment_date"],
        "UAS Assessment Type": "RA",
        "PointsAge 65": str(age_65),
        "PointsDiagnosis 3 or more coexisting Includes only documented medical diagnosis": str(diagnosis_3_plus),
        "PointsPrior history of falls within 3 months An unintentional change in position resulting in coming to rest on the ground or at a lower level": str(falls),
        "PointsIncontinence Inability to make it to the bathroom or commode in timely manner Includes frequencyurgency andor nocturia": str(incontinence),
        "PointsVisual impairment Includes but not limited to macular degeneration diabetic retinopathies visual field lossage related changes decline in visual acuity accommodation glare tolerance depth perception and night vision or not wearing prescribed glasses or having the correct prescription": str(visual_impairment),
        "PointsImpaired functional mobility May include patients who need help with IADLS or ADLS or have gait or transfer problemsarthritis pain fear of falling foot problems impaired sensation impaired coordination or improper use of assistive devices": str(impaired_functional_mobility),
        "PointsEnvironmental hazards May include but not limited to poor illumination equipment tubing inappropriate footwear pets hard to reach items floor surfaces that are uneven or cluttered or outdoor entry and exits": str(environmental_hazards),
        "PointsPoly Pharmacy 4 or more prescriptions any type All PRESCRIPTIONS including prescriptions for OTC meds Drugs highly associated with fallrisk include but not limited to sedatives antidepressants tranquilizers narcotics antihypertensives cardiac meds corticosteroids antianxiety drugs anticholinergic drugs and hypoglycemic drugs": str(polypharmacy),
        "PointsPain affecting level of function Pain often affects an individuals desire or ability to move or pain can be a factor in depression or compliance with safety recommendations": str(pain),
        "PointsCognitive impairment Could include patients with dementia Alzheimers or stroke patients or patients who are confused use poor judgment have decreased comprehension impulsivity memory deficits Consider patients ability to adhere to the plan of care": str(cognitive_impairment),
        "PointsAdd up the points and enter a Total": str(fra_total),
    }

    for page in writer.pages:
        try:
            writer.update_page_form_field_values(page, fields)
        except Exception:
            pass

    try:
        if "/AcroForm" in reader.trailer["/Root"]:
            writer._root_object.update({NameObject("/AcroForm"): reader.trailer["/Root"]["/AcroForm"]})
            writer._root_object["/AcroForm"].update({NameObject("/NeedAppearances"): BooleanObject(True)})
    except Exception:
        pass

    try:
        packet = BytesIO()
        can = canvas.Canvas(packet)
        can.setFont("Helvetica", 10)
        can.drawString(205, 47, assessor_plain)
        can.save()
        packet.seek(0)
        overlay_reader = PdfReader(packet)
        writer.pages[0].merge_page(overlay_reader.pages[0])
    except Exception:
        pass

    with open(output_path, "wb") as f:
        writer.write(f)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv", required=True)
    parser.add_argument("--case-id", required=True)
    parser.add_argument("--tt-template", default="TT_template.xlsx")
    parser.add_argument("--cdpas-template", default="CDPAS_template.xlsx")
    parser.add_argument("--fra-template", default="FRA_template.pdf")
    args = parser.parse_args()

    patient = load_patient(args.csv)
    base_name = build_base_name(patient, args.case_id)

    patient_folder = build_patient_folder(patient)
    patient_folder.mkdir(parents=True, exist_ok=True)

    tt_out = patient_folder / f"{base_name}-TT.xlsx"
    cdpas_out = patient_folder / f"{base_name}-CDPAS.xlsx"
    fra_out = patient_folder / f"{base_name}-FRA.pdf"

    fill_tt(patient, args.tt_template, tt_out)
    fill_cdpas(patient, args.cdpas_template, cdpas_out)
    fill_fra(patient, args.fra_template, fra_out, args.case_id)

    print("Generated:")
    print(tt_out)
    print(cdpas_out)
    print(fra_out)


if __name__ == "__main__":
    main()
