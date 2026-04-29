import re
from pathlib import Path
from openpyxl import load_workbook

def patch_tt_date(xlsx_path: Path, assessment_date: str):
    wb = load_workbook(xlsx_path)
    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    new_value = re.sub(
                        r'^(Date:\s*)\d{1,2}/\d{1,2}/\d{4}$',
                        r'\g<1>' + assessment_date,
                        original
                    )
                    if new_value != original:
                        cell.value = new_value
                        changed = True
    if changed:
        wb.save(xlsx_path)
