
import os
import sys

def get_resource_path(name):
    base = getattr(sys, '_MEIPASS', None)
    if base:
        return os.path.join(base, name)
    return os.path.join(os.path.dirname(__file__), name)

#!/usr/bin/env python3
import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject, BooleanObject
from tt_service_patch import apply_tt_schedule_override_v2

# === PHQ9 PATCH IMPORT START ===
from phq9_patch import should_generate_phq9_from_rows, generate_phq9_pdf
from csv_extractors import extract_member_name_file_from_rows, extract_member_id_from_rows
from commercial_profile import (
    TERMS_VERSION,
    TERMS_PATH,
    ensure_terms_files,
    load_user_profile,
    save_user_profile,
    get_assessor_display_name,
    get_assessor_plain_name,
)
from license_manager import require_active_license
# === PHQ9 PATCH IMPORT END ===


if getattr(sys, "frozen", False):
    APP_DIR = Path(sys.executable).resolve().parent
else:
    APP_DIR = Path(__file__).resolve().parent

RUNTIME_DIR = Path.home() / "Library" / "Application Support" / "UASGenerator2026V2"
RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
CONFIG_PATH = RUNTIME_DIR / "config.json"

REQUIRED_FILES = [
    "generate_outputs.py",
    "uas_automation.py",
    "TT_template.xlsx",
    "CDPAS_template.xlsx",
    "FRA_template.pdf",
]

CMVISIT_TEMPLATE_CANDIDATES = ["CM_InPerson_Visit_Template.pdf"]
PHQ9_TEMPLATE_CANDIDATES = ["PHQ9_template.pdf"]

COMPLETED_BY_NAME = get_assessor_display_name()


def run_osascript(script: str):
    result = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True
    )
    return result.returncode, result.stdout.strip(), result.stderr.strip()


def applescript_escape(text: str):
    return text.replace("\\", "\\\\").replace('"', '\\"')


def show_message(title: str, message: str):
    script = f'display dialog "{applescript_escape(message)}" with title "{applescript_escape(title)}" buttons {{"OK"}} default button "OK"'
    run_osascript(script)


def show_error(message: str):
    script = f'display dialog "{applescript_escape(message)}" with title "UAS Generator 2026" buttons {{"OK"}} default button "OK" with icon stop'
    run_osascript(script)


def choose_file_csv():
    script = '''
    try
        POSIX path of (choose file with prompt "Select CSV file" of type {"public.comma-separated-values-text", "csv"})
    on error number -128
        return "__CANCELLED__"
    end try
    '''
    code, out, _ = run_osascript(script)
    if out == "__CANCELLED__":
        return None
    return out if code == 0 and out else None


def ask_patient_string():
    script = '''
    try
        text returned of (display dialog "Enter patient string (example: First_Last_96445)" default answer "" with title "UAS Generator 2026" buttons {"Cancel", "OK"} default button "OK")
    on error number -128
        return "__CANCELLED__"
    end try
    '''
    code, out, _ = run_osascript(script)
    if out == "__CANCELLED__":
        return None
    out = out.strip()
    return out if code == 0 and out else None


def choose_output_folder():
    script = '''
    try
        POSIX path of (choose folder with prompt "Select output folder")
    on error number -128
        return "__CANCELLED__"
    end try
    '''
    code, out, _ = run_osascript(script)
    if out == "__CANCELLED__":
        return None
    return out if code == 0 and out else None


def ask_text_value(title: str, prompt: str, default_value: str = ""):
    script = f"""
    try
        text returned of (display dialog "{applescript_escape(prompt)}" default answer "{applescript_escape(default_value)}" with title "{applescript_escape(title)}" buttons {{"Cancel", "OK"}} default button "OK")
    on error number -128
        return "__CANCELLED__"
    end try
    """
    code, out, _ = run_osascript(script)
    if out == "__CANCELLED__":
        return None
    out = out.strip()
    return out if code == 0 else None


def confirm_terms_acceptance():
    subprocess.run(["open", "-a", "TextEdit", str(TERMS_PATH)], capture_output=True, text=True)
    script = f"""
    try
        button returned of (display dialog "Please review Terms_of_Use.txt in TextEdit. By clicking Accept, you confirm that this software is assistive only, that you will verify all outputs before use or signature, that no refunds are provided, and that reverse engineering, redistribution, and key sharing are prohibited." with title "Anthem BCBS Doc Generator - Terms" buttons {{"Cancel", "Decline", "Accept"}} default button "Accept")
    on error number -128
        return "__CANCELLED__"
    end try
    """
    code, out, _ = run_osascript(script)
    if out in {"__CANCELLED__", "Cancel", "Decline"}:
        return False
    return code == 0 and out == "Accept"


def ensure_user_profile_interactive(no_dialogs=False):
    ensure_terms_files()
    profile = load_user_profile()

    needs_identity = (
        not str(profile.get("first_name", "")).strip()
        or not str(profile.get("last_name", "")).strip()
        or not str(profile.get("display_name", "")).strip()
        or not str(profile.get("email", "")).strip()
    )
    needs_terms = (
        profile.get("accepted_terms_version") != TERMS_VERSION
        or not str(profile.get("accepted_at", "")).strip()
    )

    if not needs_identity and not needs_terms:
        return profile

    if no_dialogs:
        raise RuntimeError("Onboarding is incomplete. Launch from the app UI and complete setup first.")

    first_name = ask_text_value("First Launch Setup", "Enter your first name:", str(profile.get("first_name", "")))
    if not first_name:
        raise RuntimeError("Onboarding cancelled before first name was entered.")

    last_name = ask_text_value("First Launch Setup", "Enter your last name:", str(profile.get("last_name", "")))
    if not last_name:
        raise RuntimeError("Onboarding cancelled before last name was entered.")

    default_display = str(profile.get("display_name", "")).strip() or f"{first_name.strip()} {last_name.strip()} RN"
    display_name = ask_text_value("First Launch Setup", "Enter your professional display name for documents:", default_display)
    if not display_name:
        raise RuntimeError("Onboarding cancelled before display name was entered.")

    email = ask_text_value("First Launch Setup", "Enter your email address:", str(profile.get("email", "")))
    if not email:
        raise RuntimeError("Onboarding cancelled before email was entered.")

    if not confirm_terms_acceptance():
        raise RuntimeError("Terms of Use were not accepted.")

    now = datetime.now().isoformat(timespec="seconds")
    profile.update({
        "first_name": first_name.strip(),
        "last_name": last_name.strip(),
        "display_name": display_name.strip(),
        "email": email.strip(),
        "accepted_terms_version": TERMS_VERSION,
        "accepted_at": now,
        "updated_at": now,
    })
    save_user_profile(profile)
    return profile


def load_config():
    if not CONFIG_PATH.exists():
        return {}
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(data: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def ensure_required_files():
    missing = [name for name in REQUIRED_FILES if not (APP_DIR / name).exists()]
    if missing:
        raise RuntimeError("Missing required files:\n" + "\n".join(missing))


def get_output_root(output_override=None, no_dialogs=False):
    config = load_config()

    if output_override:
        output_root = Path(output_override).expanduser().resolve()
        config["output_folder"] = str(output_root)
        save_config(config)
        return output_root

    saved = config.get("output_folder", "").strip()
    if saved and Path(saved).exists():
        return Path(saved)

    if no_dialogs:
        raise RuntimeError("Output folder is not set. Choose one in the app or save it first.")

    chosen = choose_output_folder()
    if not chosen:
        raise RuntimeError("Output folder selection cancelled.")

    output_root = Path(chosen).expanduser().resolve()
    config["output_folder"] = str(output_root)
    save_config(config)
    return output_root


def split_patient_string(patient_string: str):
    parts = [p.strip() for p in patient_string.split("_") if p.strip()]
    first = parts[0].upper() if len(parts) >= 1 else ""
    last = parts[1].upper() if len(parts) >= 2 else ""
    id_digits = ""
    if parts:
        id_digits = "".join(ch for ch in parts[-1] if ch.isdigit())
    if not id_digits:
        id_digits = "".join(ch for ch in patient_string if ch.isdigit())
    return first, last, id_digits


def extract_numeric_case_id(patient_string: str):
    _, _, id_digits = split_patient_string(patient_string)
    return id_digits or patient_string.strip()


def normalize_date_for_folder(mmddyyyy: str):
    try:
        dt = datetime.strptime(mmddyyyy, "%m/%d/%Y")
        return dt.strftime("%m-%d-%Y")
    except Exception:
        return mmddyyyy.replace("/", "-")


def extract_csv_rows(csv_path: Path):
    with open(csv_path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        return list(csv.reader(f))


def extract_assessment_date_from_rows(rows):
    patterns = [
        re.compile(r'Assessment Date:\s*(\d{2}/\d{2}/\d{4})', re.IGNORECASE),
        re.compile(r'Assessment Reference Date:\s*(\d{2}/\d{2}/\d{4})', re.IGNORECASE),
    ]

    for row in rows:
        joined = " | ".join(cell.strip() for cell in row if cell and cell.strip())
        for pat in patterns:
            m = pat.search(joined)
            if m:
                return m.group(1)

    for row in rows:
        for i, cell in enumerate(row):
            value = str(cell).strip().lower()
            if value in {"assessment date:", "assessment reference date:"}:
                for j in range(i + 1, len(row)):
                    val = str(row[j]).strip()
                    if re.fullmatch(r'\d{2}/\d{2}/\d{4}', val):
                        return val

    raise RuntimeError("Could not extract Assessment Date from CSV.")


def extract_mode_of_assessment_from_rows(rows):
    for idx, row in enumerate(rows):
        cleaned = [str(c).strip() for c in row if str(c).strip()]
        joined = " | ".join(cleaned).lower()

        if "mode of assessment" in joined:
            window_parts = [joined]
            for extra in range(1, 5):
                if idx + extra < len(rows):
                    next_joined = " | ".join(str(c).strip() for c in rows[idx + extra] if str(c).strip()).lower()
                    if next_joined:
                        window_parts.append(next_joined)
            window = " | ".join(window_parts)

            if "in-person only" in window:
                return "in_person"
            if "interactive video teleconference only" in window:
                return "telehealth"

    full_text = "\n".join(" | ".join(str(c).strip() for c in row if str(c).strip()).lower() for row in rows)

    if re.search(r'mode of assessment.*in-person only', full_text, re.IGNORECASE):
        return "in_person"
    if re.search(r'mode of assessment.*interactive video teleconference only', full_text, re.IGNORECASE):
        return "telehealth"

    return "telehealth"


def extract_medicaid_from_rows(rows):
    code_pat = re.compile(r'\b[A-Z]{2}\d{5,6}[A-Z]\b')
    priority_labels = [
        "member cin", "member cin no", "cin", "cin no",
        "medicaid", "medicaid number", "medicaid no",
    ]

    for row in rows:
        for i, cell in enumerate(row):
            text = str(cell).strip()
            low = text.lower()
            if any(label in low for label in priority_labels):
                scan = row[i:i+6]
                joined = " | ".join(str(x).strip() for x in scan if str(x).strip())
                m = code_pat.search(joined.upper())
                if m:
                    return m.group(0)

    for row in rows:
        joined = " | ".join(str(x).strip() for x in row if str(x).strip())
        m = code_pat.search(joined.upper())
        if m:
            return m.group(0)

    return None


def build_patient_folder_name(patient_name_file: str, assessment_date: str):
    base = (patient_name_file or "UNKNOWN_MEMBER").strip().upper()
    return f"{base} - {normalize_date_for_folder(assessment_date)}"


def parse_generated_paths(stdout_text: str):
    paths = []
    for line in stdout_text.splitlines():
        line = line.strip()
        if line.startswith("/Users/") and (line.endswith(".xlsx") or line.endswith(".pdf")):
            paths.append(Path(line))
    return paths


def classify_output(path: Path):
    name = path.name.lower()
    if name.endswith(".pdf"):
        return "FRA"
    if "cdpas" in name:
        return "CDPAS"
    if name.endswith(".xlsx"):
        return "TT"
    return None


def extract_medicaid_from_name(path: Path):
    m = re.search(r'ANTHEM_MLTC-([A-Z0-9]+)-', path.name, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return None


def remove_previous_generated_files(patient_dir: Path):
    for p in patient_dir.iterdir():
        if not p.is_file():
            continue
        name = p.name.upper()
        if (
            name.endswith("-TT.XLSX")
            or name.endswith("-CDPAS.XLSX")
            or name.endswith("-FRA.PDF")
            or name.endswith("-6MONTHCMVISIT.PDF")
            or name.endswith("-PHQ9.PDF")
            or name in {"TT.XLSX", "CDPAS.XLSX", "FRA.PDF", "PHQ9.PDF"}
        ):
            try:
                p.unlink()
            except Exception:
                pass


def final_filename(medicaid_id: str, last: str, first: str, numeric_id: str, doc_type: str):
    base = f"ANTHEM_MLTC-{medicaid_id}-{last}_{first}-{numeric_id}-{doc_type}"
    if doc_type in {"FRA", "6monthCMVisit", "PHQ9"}:
        return base + ".pdf"
    return base + ".xlsx"


def patch_cdpas_date(xlsx_path: Path, assessment_date: str):
    wb = load_workbook(xlsx_path)
    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    new_value = re.sub(
                        r'(Date of Plan:\s*)\d{1,2}/\d{1,2}/\d{4}',
                        r'\g<1>' + assessment_date,
                        original
                    )
                    if new_value != original:
                        cell.value = new_value
                        changed = True
    if changed:
        wb.save(xlsx_path)


def patch_tt_date(xlsx_path: Path, assessment_date: str):
    wb = load_workbook(xlsx_path)
    changed = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    new_value = re.sub(
                        r'^(Date:\s*)\d{1,2}/\d{1,2}/\d{4}$',
                        r'\g<1>' + assessment_date,
                        original
                    )
                    if new_value != original:
                        cell.value = new_value
                        changed = True
    if changed:
        wb.save(xlsx_path)



def export_excel_to_pdf(xlsx_path: Path):
    pdf_path = xlsx_path.with_suffix(".pdf")

    # First try Microsoft Excel on Mac
    excel_app = Path("/Applications/Microsoft Excel.app")
    if excel_app.exists():
        script = f"""
        tell application "Microsoft Excel"
            activate
            open POSIX file "{xlsx_path}"
            delay 1
            save active workbook in POSIX file "{pdf_path}" as PDF file format
            close active workbook saving no
        end tell
        """
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True
        )
        if result.returncode == 0 and pdf_path.exists():
            return

    # Fallback: LibreOffice / soffice
    soffice_candidates = [
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]

    for candidate in soffice_candidates:
        try:
            result = subprocess.run(
                [candidate, "--headless", "--convert-to", "pdf", "--outdir", str(xlsx_path.parent), str(xlsx_path)],
                capture_output=True,
                text=True
            )
            if result.returncode == 0 and pdf_path.exists():
                return
        except FileNotFoundError:
            continue

    raise RuntimeError(
        f"Could not export PDF for {xlsx_path.name}. "
        f"Install Microsoft Excel or LibreOffice, or tell me and I will add a different export method."
    )



def _tt_num(value):
    try:
        if value is None or value == "":
            return 0
        return float(value)
    except Exception:
        return 0

def _tt_recalc_sheet(ws):
    """
    Keep TT formulas alive.

    The program still fills CSV-derived input cells in B:H and I:J.
    This function must NOT replace formula cells with hardcoded numbers.
    K12:K26 and K27 stay as Excel formulas so manual edits recalculate.
    """
    for row in range(12, 27):
        ws[f"K{row}"] = f"=SUM(B{row}:H{row})*I{row}*J{row}"

    ws["K27"] = "=SUM(K12:K26)/60"

    try:
        wb = ws.parent
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass






def apply_tt_schedule_override(xlsx_path: Path, days_per_week: Optional[int], hours_per_day: Optional[int]):
    if days_per_week is None and hours_per_day is None:
        return

    wb = load_workbook(xlsx_path)
    ws = wb.active

    if hours_per_day is not None:
        ws["B32"] = hours_per_day

    if days_per_week is not None:
        ws["K32"] = days_per_week

    if days_per_week is not None and hours_per_day is not None:
        ws["A36"] = f"PCA {days_per_week} D x {hours_per_day} H"
    elif days_per_week is not None:
        current_hours = ws["B32"].value
        if current_hours not in (None, ""):
            ws["A36"] = f"PCA {days_per_week} D x {current_hours} H"

    if days_per_week is not None:
        rows_to_override = [12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24]
        for row in rows_to_override:
            ws[f"J{row}"] = days_per_week
        ws["J14"] = 3
        ws["J16"] = 2
        ws["J25"] = 2
        ws["J26"] = 3
    _tt_recalc_sheet(ws)

    wb.save(xlsx_path)


def row_triggered_in_section_k(rows, label: str) -> bool:
    label = label.lower()
    for row in rows:
        row_text = " | ".join(str(c).strip() for c in row if str(c).strip())
        low = row_text.lower()
        if label in low:
            if "no selection" in low or "not ordered and did not occur" in low:
                if ("ordered, not implemented" in low or
                    "1-2 days of last 3 days" in low or
                    "daily in last 3 days" in low):
                    # In some exports all choices may appear in one row; fall through to explicit positive check
                    pass
                else:
                    return False
            if ("ordered, not implemented" in low or
                "1-2 days of last 3 days" in low or
                "daily in last 3 days" in low):
                return True
    return False


def contains_text(rows, terms):
    joined = "\n".join(" | ".join(str(c).strip() for c in row if str(c).strip()) for row in rows).lower()
    return any(term.lower() in joined for term in terms)




def parse_csv_cdpas_flags(rows):
    wound_care = False
    insulin = False

    positive_k_values = {
        "Ordered, not implemented",
        "1-2 days of last 3 days",
        "Daily in last 3 days",
    }

    for row in rows:
        parts = [str(c).strip() for c in row if str(c).strip()]
        if not parts:
            continue
        joined = " | ".join(parts)
        low = joined.lower()

        if "wound care" in low:
            for v in positive_k_values:
                if v.lower() in low:
                    wound_care = True

        if "insulin" in low or "subcutaneous injection" in low or "subcutaneous injections" in low or "sq injection" in low:
            insulin = True

    return {
        "wound_care": wound_care,
        "insulin": insulin,
    }


def parse_rom_value_from_equipment(equipment):
    if equipment.get("wheelchair") or equipment.get("hospital_bed") or equipment.get("hoyer_lift"):
        return "Passive"
    if equipment.get("cane") or equipment.get("walker"):
        return "Active"
    return ""


def parse_continence_from_rows(rows):
    bladder_val = ""
    bowel_val = ""

    for row in rows:
        parts = [str(c).strip() for c in row if str(c).strip()]
        if not parts:
            continue
        joined = " | ".join(parts)
        low = joined.lower()

        if "bladder continence" in low and not bladder_val:
            if "|" in joined:
                bladder_val = joined.split("|")[-1].strip().lower()
            else:
                bladder_val = low

        if "bowel continence" in low and not bowel_val:
            if "|" in joined:
                bowel_val = joined.split("|")[-1].strip().lower()
            else:
                bowel_val = low

    incont_terms = ["incontinent", "usually incontinent", "occasionally incontinent"]
    cont_terms = ["continent", "completely continent"]

    bladder_incont = any(t in bladder_val for t in incont_terms)
    bowel_incont = any(t in bowel_val for t in incont_terms)

    bladder_cont = any(t in bladder_val for t in cont_terms)
    bowel_cont = any(t in bowel_val for t in cont_terms)

    if bladder_incont or bowel_incont:
        return "incontinent"

    if bladder_cont and bowel_cont:
        return "continent"

    return "unknown"


def parse_reason_for_assessment(rows):
    parts = []
    for row in rows:
        try:
            iterable = list(row)
        except Exception:
            iterable = [row]
        for cell in iterable:
            if cell is None:
                continue
            s = str(cell).strip()
            if s:
                parts.append(s.lower())

    blob = " | ".join(parts)

    if "significant change in status reassessment" in blob:
        return "scic"
    if "routine reassessment" in blob:
        return "routine"

    if "reason for assessment" in blob and "other" in blob:
        return "scic"

    return None




def find_cmvisit_template():
    for name in CMVISIT_TEMPLATE_CANDIDATES:
        path = APP_DIR / name
        if path.exists():
            return path
    for path in APP_DIR.glob("*6monthCMVisit.pdf"):
        if path.is_file():
            return path
    return None


def create_cmvisit_pdf(
    template_path: Path,
    output_path: Path,
    first: str,
    last: str,
    numeric_id: str,
    medicaid_id: str,
    assessment_date: str,
    equipment=None,
    dwelling_type="apartment_elevator",
    bedrooms_count="2",
    accessible_home="no",
    member_name_value=None,
):
    from cmvisit import create_cmvisit_pdf as _impl
    return _impl(
        template_path,
        output_path,
        first,
        last,
        numeric_id,
        medicaid_id,
        assessment_date,
        equipment=equipment,
        dwelling_type=dwelling_type,
        bedrooms_count=bedrooms_count,
        accessible_home=accessible_home,
        member_name_value=member_name_value,
    )

def extract_dob_from_rows(rows):
    patterns = [
        re.compile(r"Date of Birth:\s*(\d{2}/\d{2}/\d{4})", re.IGNORECASE),
        re.compile(r"DOB:\s*(\d{2}/\d{2}/\d{4})", re.IGNORECASE),
    ]

    for row in rows:
        joined = " | ".join(str(cell).strip() for cell in row if str(cell).strip())
        for pat in patterns:
            m = pat.search(joined)
            if m:
                return m.group(1)

    for row in rows:
        for i, cell in enumerate(row):
            value = str(cell).strip().lower()
            if value in {"date of birth:", "dob:"}:
                for j in range(i + 1, len(row)):
                    val = str(row[j]).strip()
                    if re.fullmatch(r"\d{2}/\d{2}/\d{4}", val):
                        return val

    return ""




def create_cdpas_via_excel(template_path: Path, output_path: Path, assessment_date: str, rows, equipment, medicaid_id: str, member_name_value: str):
    if not Path("/Applications/Microsoft Excel.app").exists():
        raise RuntimeError("Microsoft Excel is required for CDPAS generation with logo preservation.")

    shutil.copy2(template_path, output_path)

    dob_value = extract_dob_from_rows(rows)
    kind = parse_reason_for_assessment(rows)
    flags = parse_csv_cdpas_flags(rows)
    continence = parse_continence_from_rows(rows)

    if not isinstance(flags, dict):
        flags = {}

    wound_care_flag = bool(flags.get("wound_care"))
    insulin_flag = bool(flags.get("insulin"))

    aspiration_flag = bool(equipment.get("cdpas_aspiration_precautions")) or bool(equipment.get("cdpas_tube_feeding"))
    tube_feeding_flag = bool(equipment.get("cdpas_tube_feeding"))
    oxygen_care_flag = bool(equipment.get("cdpas_oxygen_care"))
    catheter_type = str(equipment.get("cdpas_catheter_type") or "").strip()
    catheter_flag = bool(catheter_type)

    mobility_items = []
    for key, label in [("cane", "Cane"), ("walker", "Walker"), ("wheelchair", "Wheelchair")]:
        if equipment.get(key):
            mobility_items.append(label)

    dme_items = []
    for key, label in [("hospital_bed", "Hospital bed"), ("hoyer_lift", "Hoyer lift")]:
        if equipment.get(key):
            dme_items.append(label)

    rom_value = parse_rom_value_from_equipment(equipment)

    cell_values = {
        "B1": member_name_value,
        "C1": f"DOB: {dob_value}" if dob_value else "",
        "E1": f"CIN: {medicaid_id}",
        "E2": f"Date of Plan: {assessment_date}",
        "E4": f"RN Signature: {get_assessor_display_name()}",
        "G1": "Initial",
        "G2": "SCIC",
        "G3": "Other",
        "H1": "Routine/Reassessment",
        "H2": "Return",

        "C27": mobility_items[0] if len(mobility_items) > 0 else "",
        "D27": mobility_items[1] if len(mobility_items) > 1 else "",

        "C28": dme_items[0] if len(dme_items) > 0 else "",
        "D28": dme_items[1] if len(dme_items) > 1 else "",

        "C29": rom_value,

        "G8": "X" if aspiration_flag else "",
        "G9": "X" if tube_feeding_flag else "",
        "G10": "X" if tube_feeding_flag else "",

        "G11": "X" if insulin_flag else "",
        "G12": "X" if insulin_flag else "",
        "G13": "X" if insulin_flag else "",
        "G15": "X" if insulin_flag else "",

        "G17": "X" if wound_care_flag else "",

        "G19": "X",
        "G20": "X",
        "G21": "X",
        "G22": "X" if continence == "incontinent" else "",
        "G23": "X" if continence == "incontinent" else "",

        "G26": catheter_type if catheter_flag else "",
        "G30": "X" if oxygen_care_flag else "",
        "G35": "X" if wound_care_flag else "",
    }
    if kind == "routine":
        cell_values["G1"] = "Initial"
        cell_values["G2"] = "SCIC"
        cell_values["G3"] = "Other"
        cell_values["H1"] = "Routine/Reassessment          X"
        cell_values["H2"] = "Return"
    elif kind == "scic":
        cell_values["G1"] = "Initial"
        cell_values["G2"] = "SCIC   X"
        cell_values["G3"] = "Other"
        cell_values["H1"] = "Routine/Reassessment"
        cell_values["H2"] = "Return"
    else:
        cell_values["G1"] = "Initial"
        cell_values["G2"] = "SCIC"
        cell_values["G3"] = "Other"
        cell_values["H1"] = "Routine/Reassessment"
        cell_values["H2"] = "Return"
    set_lines = []
    for cell_ref, value in cell_values.items():
        safe = applescript_escape("" if value is None else str(value))
        set_lines.append(f'set value of range "{cell_ref}" of active sheet to "{safe}"')

    script_lines = [
        'try',
        '    tell application "Microsoft Excel"',
        '        activate',
        f'        open POSIX file "{applescript_escape(str(output_path))}"',
        '        delay 4',
        '        repeat 20 times',
        '            try',
        '                set wbName to name of active workbook',
        '                exit repeat',
        '            on error',
        '                delay 1',
        '            end try',
        '        end repeat',
    ]
    script_lines.extend(['        ' + line for line in set_lines])
    script_lines.extend([
        '        delay 2',
        '        save active workbook',
        '        delay 2',
        '        try',
        '            close active workbook saving yes',
        '        end try',
        '    end tell',
        '    return "OK"',
        'on error errMsg',
        '    return "FAIL: " & errMsg',
        'end try',
    ])

    script = "\n".join(script_lines)

    code, out, err = run_osascript(script)

    if code == 0 and out.startswith("OK"):
        return

    raise RuntimeError(f"Excel CDPAS write failed: {out or err}")


def copy_outputs_to_patient_dir(
    generated_paths,
    patient_dir: Path,
    patient_name_file: str,
    numeric_id: str,
    assessment_date: str,
    mode: str,
    medicaid_id: str,
    days_per_week: Optional[int],
    hours_per_day: Optional[int],
    rows,
    equipment,
    dwelling_type="apartment_elevator",
    service_flag=False,
    bedrooms_count="2",
    accessible_home="no",
):
    copied = {}
    patient_dir.mkdir(parents=True, exist_ok=True)
    remove_previous_generated_files(patient_dir)

    csv_name = (patient_name_file or "").strip().upper()
    parts = [p.strip() for p in csv_name.split("_") if p.strip()]
    if len(parts) >= 2:
        last = parts[0]
        first = "_".join(parts[1:])
    elif len(parts) == 1:
        last = parts[0]
        first = ""
    else:
        raise RuntimeError("Could not extract patient name from CSV.")

    if not numeric_id:
        raise RuntimeError("Could not extract Member ID from CSV or UI.")

    if not medicaid_id:
        for src in generated_paths:
            found = extract_medicaid_from_name(src)
            if found:
                medicaid_id = found
                break

    if not medicaid_id:
        raise RuntimeError("Could not extract Medicaid/CIN number from CSV or generated filenames.")

    for src in generated_paths:
        if not src.exists():
            continue

        doc_type = classify_output(src)
        if not doc_type:
            continue

        dest = patient_dir / final_filename(medicaid_id, last, first, numeric_id, doc_type)

        if doc_type == "CDPAS":
            create_cdpas_via_excel(
                APP_DIR / "CDPAS_template.xlsx",
                dest,
                assessment_date,
                rows,
                equipment,
                medicaid_id,
                csv_name,
            )
            export_excel_to_pdf(dest)
            copied[doc_type] = dest
            continue

        shutil.copy2(src, dest)

        if doc_type == "TT":
            patch_tt_date(dest, assessment_date)
            apply_tt_schedule_override_v2(dest, days_per_week, hours_per_day, rows, service_flag=service_flag)
            export_excel_to_pdf(dest)

        copied[doc_type] = dest

    if mode == "in_person":
        template_path = find_cmvisit_template()
        if not template_path:
            raise RuntimeError("In-person assessment detected, but CM visit template PDF was not found in the project folder.")

        cmvisit_dest = patient_dir / final_filename(medicaid_id, last, first, numeric_id, "6monthCMVisit")
        create_cmvisit_pdf(
            template_path,
            cmvisit_dest,
            first,
            last,
            numeric_id,
            medicaid_id,
            assessment_date,
            equipment=equipment,
            dwelling_type=dwelling_type,
            bedrooms_count=bedrooms_count,
            accessible_home=accessible_home,
            member_name_value=csv_name,
        )
        copied["6monthCMVisit"] = cmvisit_dest

    return copied, medicaid_id, numeric_id


def run_generator(
    csv_path: Path,
    patient_name_file: str,
    numeric_id: str,
    patient_dir: Path,
    assessment_date: str,
    mode: str,
    medicaid_id: str,
    days_per_week: Optional[int],
    hours_per_day: Optional[int],
    rows,
    equipment,
    dwelling_type="apartment_elevator",
    service_flag=False,
    bedrooms_count="2",
    accessible_home="no",
):
    generator = APP_DIR / "generate_outputs.py"
    python_bin = sys.executable
    case_id = str(numeric_id).strip()

    env = os.environ.copy()
    env["PYTHONPATH"] = str(APP_DIR) + (os.pathsep + env["PYTHONPATH"] if env.get("PYTHONPATH") else "")
    env["ASSESSMENT_DATE"] = assessment_date
    env["MODE_OF_ASSESSMENT"] = mode
    env["ASSESSOR_DISPLAY_NAME"] = get_assessor_display_name()
    env["ASSESSOR_PLAIN_NAME"] = get_assessor_plain_name()
    if medicaid_id:
        env["MEDICAID_ID"] = medicaid_id

    cmd = [
        python_bin,
        str(generator),
        "--csv", str(csv_path),
        "--case-id", case_id,
        "--tt-template", str(APP_DIR / "TT_template.xlsx"),
        "--cdpas-template", str(APP_DIR / "CDPAS_template.xlsx"),
        "--fra-template", str(APP_DIR / "FRA_template.pdf"),
    ]

    result = subprocess.run(
        cmd,
        cwd=str(APP_DIR),
        env=env,
        capture_output=True,
        text=True
    )

    stdout_text = result.stdout.strip()
    stderr_text = result.stderr.strip()

    generated_paths = parse_generated_paths(stdout_text)
    copied, medicaid_id, numeric_id = copy_outputs_to_patient_dir(
        generated_paths,
        patient_dir,
        patient_name_file,
        numeric_id,
        assessment_date,
        mode,
        medicaid_id,
        days_per_week,
        hours_per_day,
        rows,
        equipment,
        dwelling_type,
        service_flag=service_flag,
        bedrooms_count=bedrooms_count,
        accessible_home=accessible_home,
    )

    required = ["TT", "CDPAS", "FRA"]
    if mode == "in_person":
        required.append("6monthCMVisit")

    success = all(k in copied for k in required)
    return success, stdout_text, stderr_text, copied, case_id, medicaid_id, numeric_id


def parse_optional_int(value, field_name):
    if value is None or value == "":
        return None
    try:
        n = int(str(value).strip())
    except Exception:
        raise RuntimeError(f"{field_name} must be a whole number.")
    if n <= 0:
        raise RuntimeError(f"{field_name} must be greater than 0.")
    return n


def perform_generation(
    csv_path_str: str,
    patient_string: str = None,
    output_override=None,
    days_per_week=None,
    hours_per_day=None,
    equipment=None,
    dwelling_type="apartment_elevator",
    service_flag=False,
    bedrooms_count=None,
    accessible_home="no",
):
    require_active_license()
    ensure_required_files()

    if equipment is None:
        equipment = {
            "cane": False,
            "walker": False,
            "wheelchair": False,
            "hospital_bed": False,
            "catheter": False,
            "grab_bar": False,
            "bedside_commode": False,
            "oxygen": False,
            "service": False,
            "hoyer_lift": False,
            "raised_toilet_seat": False,
            "shower_chair": False,
            "cdpas_aspiration_precautions": False,
            "cdpas_tube_feeding": False,
            "cdpas_oxygen_care": False,
            "cdpas_catheter_type": "",
        }

    selected_mobility_count = sum(1 for k in ["cane", "walker", "wheelchair"] if equipment.get(k))
    if selected_mobility_count > 2:
        raise RuntimeError("Choose at most 2 of Cane, Walker, Wheelchair.")

    days_per_week = parse_optional_int(days_per_week, "Days per week")
    hours_per_day = parse_optional_int(hours_per_day, "Hours per day")

    if bedrooms_count is None or str(bedrooms_count).strip() == "":
        bedrooms_count = "2"
    else:
        bedrooms_count = str(bedrooms_count).strip()
        try:
            if int(bedrooms_count) < 0:
                raise RuntimeError("Bedrooms count must be 0 or greater.")
        except ValueError:
            raise RuntimeError("Bedrooms count must be a whole number.")

    accessible_home = str(accessible_home or "no").strip().lower()
    if accessible_home not in {"yes", "no"}:
        accessible_home = "no"

    csv_path = Path(csv_path_str).expanduser().resolve()
    rows = extract_csv_rows(csv_path)
    assessment_date = extract_assessment_date_from_rows(rows)
    mode = extract_mode_of_assessment_from_rows(rows)
    medicaid_id = extract_medicaid_from_rows(rows)

    patient_name_file = extract_member_name_file_from_rows(rows)
    numeric_id = extract_member_id_from_rows(rows)

    if not patient_name_file:
        raise RuntimeError("Could not extract patient name from CSV.")

    override_id = extract_numeric_case_id(patient_string) if patient_string else ""
    if override_id:
        numeric_id = override_id

    if not numeric_id:
        raise RuntimeError("Could not extract Member ID from CSV, and no Patient ID was entered.")

    output_root = get_output_root(output_override=output_override, no_dialogs=True)
    output_root.mkdir(parents=True, exist_ok=True)

    patient_folder_name = build_patient_folder_name(patient_name_file, assessment_date)
    patient_dir = output_root / patient_folder_name
    patient_dir.mkdir(parents=True, exist_ok=True)

    success, stdout_text, stderr_text, copied, case_id, medicaid_id, numeric_id = run_generator(
        csv_path,
        patient_name_file,
        numeric_id,
        patient_dir,
        assessment_date,
        mode,
        medicaid_id,
        days_per_week,
        hours_per_day,
        rows,
        equipment,
        dwelling_type,
        service_flag=service_flag,
        bedrooms_count=bedrooms_count,
        accessible_home=accessible_home,
    )

    phq_created = None
    if should_generate_phq9_from_rows(rows):
        phq_template_path = APP_DIR / "PHQ9_template.pdf"
        if phq_template_path.exists():
            parts = [p.strip() for p in patient_name_file.strip().upper().split("_") if p.strip()]
            if len(parts) >= 2:
                last = parts[0]
                first = "_".join(parts[1:])
            elif len(parts) == 1:
                last = parts[0]
                first = ""
            else:
                last = "UNKNOWN"
                first = ""
            phq_dest = patient_dir / final_filename(medicaid_id, last, first, numeric_id, "PHQ9")
            generate_phq9_pdf(
                phq_template_path,
                phq_dest,
                patient_name_file,
                numeric_id,
                assessment_date,
            )
            copied["PHQ9"] = phq_dest
            phq_created = phq_dest

    if not success:
        details = []
        if stdout_text:
            details.append("STDOUT:\n" + stdout_text)
        if stderr_text:
            details.append("STDERR:\n" + stderr_text)
        message = (
            "Generation did not complete.\n\n"
            f"Patient folder:\n{patient_dir}\n\n"
            f"Assessment Date extracted:\n{assessment_date}\n\n"
            f"Mode extracted:\n{mode}\n\n"
            f"Medicaid extracted:\n{medicaid_id}\n\n"
            f"Case ID extracted:\n{case_id}\n\n"
            f"Patient name extracted:\n{patient_name_file}"
        )
        if details:
            message += "\n\n" + "\n\n".join(details[:2])
        raise RuntimeError(message)

    lines = [
        "Files created successfully:",
        str(copied["TT"]),
        str(copied["CDPAS"]),
        str(copied["FRA"]),
    ]
    if "6monthCMVisit" in copied:
        lines.append(str(copied["6monthCMVisit"]))
    if "PHQ9" in copied:
        lines.append(str(copied["PHQ9"]))
    lines.extend([
        "",
        f"Assessment Date used: {assessment_date}",
        f"Mode: {mode}",
        f"Medicaid: {medicaid_id}",
        f"Case ID: {numeric_id}",
        f"Patient name: {patient_name_file}",
        f"Bedrooms count: {bedrooms_count}",
        f"Accessible home: {accessible_home}",
        f"Equipment: {equipment}",
    ])
    if days_per_week is not None:
        lines.append(f"Days per week override: {days_per_week}")
    if hours_per_day is not None:
        lines.append(f"Hours per day override: {hours_per_day}")
    lines.append(f"Dwelling type: {dwelling_type}")
    lines.append("Source of truth for patient name: CSV")
    lines.append("Source of truth for Member ID: Patient ID field if entered, otherwise CSV")
    return "\n".join(lines)


def main():
    ensure_terms_files()
    parser = argparse.ArgumentParser()
    parser.add_argument("--csv")
    parser.add_argument("--patient")
    parser.add_argument("--output")
    parser.add_argument("--days-per-week")
    parser.add_argument("--hours-per-day")
    parser.add_argument("--dwelling-type", default="apartment_elevator")
    parser.add_argument("--bedrooms-count")
    parser.add_argument("--accessible-home", default="no")
    parser.add_argument("--cane", action="store_true")
    parser.add_argument("--walker", action="store_true")
    parser.add_argument("--wheelchair", action="store_true")
    parser.add_argument("--hospital-bed", action="store_true")
    parser.add_argument("--catheter", action="store_true")
    parser.add_argument("--grab-bar", action="store_true")
    parser.add_argument("--bedside-commode", action="store_true")
    parser.add_argument("--oxygen-in-use", action="store_true")
    parser.add_argument("--service", action="store_true")
    parser.add_argument("--hoyer-lift", action="store_true")
    parser.add_argument("--raised-toilet-seat", action="store_true")
    parser.add_argument("--shower-chair", action="store_true")
    parser.add_argument("--cdpas-aspiration-precautions", action="store_true")
    parser.add_argument("--cdpas-tube-feeding", action="store_true")
    parser.add_argument("--cdpas-oxygen-care", action="store_true")
    parser.add_argument("--cdpas-catheter-type")
    parser.add_argument("--no-dialogs", action="store_true")
    args = parser.parse_args()

    equipment = {
        "cane": args.cane,
        "walker": args.walker,
        "wheelchair": args.wheelchair,
        "hospital_bed": args.hospital_bed,
        "catheter": args.catheter,
        "grab_bar": args.grab_bar,
        "bedside_commode": args.bedside_commode,
        "oxygen": args.oxygen_in_use,
        "service": args.service,
        "hoyer_lift": args.hoyer_lift,
        "raised_toilet_seat": args.raised_toilet_seat,
        "shower_chair": args.shower_chair,
        "cdpas_aspiration_precautions": args.cdpas_aspiration_precautions,
        "cdpas_tube_feeding": args.cdpas_tube_feeding,
        "cdpas_oxygen_care": args.cdpas_oxygen_care,
        "cdpas_catheter_type": args.cdpas_catheter_type or "",
    }

    cli_mode = bool(args.csv)

    try:
        if cli_mode:
            result = perform_generation(
                csv_path_str=args.csv,
                patient_string=args.patient,
                output_override=args.output,
                days_per_week=args.days_per_week,
                hours_per_day=args.hours_per_day,
                equipment=equipment,
                dwelling_type=args.dwelling_type,
                service_flag=args.service,
                bedrooms_count=args.bedrooms_count,
                accessible_home=args.accessible_home,
            )
            if args.no_dialogs:
                print(result)
            else:
                show_message("UAS Generator 2026", result)
            sys.exit(0)

        ensure_required_files()

        csv_selected = choose_file_csv()
        if not csv_selected:
            sys.exit(0)

        patient_string = ask_patient_string()
        if not patient_string:
            sys.exit(0)

        config = load_config()
        saved = config.get("output_folder", "").strip()
        if not saved or not Path(saved).exists():
            chosen = choose_output_folder()
            if not chosen:
                sys.exit(0)
            config["output_folder"] = str(Path(chosen).expanduser().resolve())
            save_config(config)

        result = perform_generation(
            csv_path_str=csv_selected,
            patient_string=patient_string,
            output_override=None,
            days_per_week=None,
            hours_per_day=None,
            equipment=equipment,
            dwelling_type="apartment_elevator",
            service_flag=False,
            bedrooms_count="2",
            accessible_home="no",
        )
        show_message("UAS Generator 2026", result)
        sys.exit(0)

    except Exception as e:
        if args.no_dialogs:
            print(str(e), file=sys.stderr)
        else:
            show_error(str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()
